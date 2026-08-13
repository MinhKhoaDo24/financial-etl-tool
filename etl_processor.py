import re
import io
import xlrd
import openpyxl
import pdfplumber
import pandas as pd

from merge_processor import format_customer_name, is_company_or_organization


# Mapping between sheet names in PBSV-style multi-sheet Excel → canonical table names
PBSV_SHEET_MAP = {
    'CoPhieu':          'co_phieu',
    'CongTyChungKhoan': 'cong_ty_chung_khoan',
    'NguoiQuanLy':      'nguoi_quan_ly',
    'ChinhSach':        'chinh_sach',
    'PhanLoaiKhachHang':'phan_loai_khach_hang',
    'NhomKhachHang':    'nhom_khach_hang',
    'KhachHang':        'khach_hang',
    'GiaoDich':         'giao_dich',
    'PhiGiaHan':        'phi_gia_han',
    'BaoCaoThuLai':     'bao_cao_thu_lai',
}


def is_multi_sheet_data_model(file_bytes_or_path):
    """
    Returns True if the uploaded xlsx file is a PBSV-style multi-sheet Data Model workbook.
    Detection: at least 4 of the known sheet names match PBSV_SHEET_MAP keys.
    A workbook merely having >1 sheet is NOT enough — real transaction report exports
    (e.g. PBSV.xlsx "Lịch sử đặt lệnh") are single-sheet and must NOT be routed here.
    """
    try:
        if isinstance(file_bytes_or_path, bytes):
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes_or_path), data_only=True, read_only=True)
        else:
            wb = openpyxl.load_workbook(file_bytes_or_path, data_only=True, read_only=True)
        matches = sum(1 for s in wb.sheetnames if s in PBSV_SHEET_MAP)
        return matches >= 4
    except Exception:
        return False


def parse_multi_sheet_excel(file_bytes_or_path, filename=""):
    """
    Reads a PBSV-style multi-sheet Data Model workbook and returns a dictionary
    mapping canonical table names -> pandas DataFrames, ready to use as db_tables.
    """
    if isinstance(file_bytes_or_path, bytes):
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes_or_path), data_only=True)
    else:
        wb = openpyxl.load_workbook(file_bytes_or_path, data_only=True)

    db_tables = {}
    for sheet_name, canonical_name in PBSV_SHEET_MAP.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c).strip() if c is not None else f"Col{i}" for i, c in enumerate(rows[0])]
        data_rows = []
        for row in rows[1:]:
            if all(v is None for v in row):
                continue
            data_rows.append(list(row))
        df = pd.DataFrame(data_rows, columns=header)
        for col in df.columns:
            if df[col].dtype == object:
                df[col] = df[col].apply(lambda v: v.strftime('%d/%m/%Y') if hasattr(v, 'strftime') else v)
        db_tables[canonical_name] = df

    for canonical_name in PBSV_SHEET_MAP.values():
        if canonical_name not in db_tables:
            db_tables[canonical_name] = pd.DataFrame()

    return db_tables


# ────────────────────────────────────────────────────────────────────────────
# Shared low-level helpers
# ────────────────────────────────────────────────────────────────────────────

def _safe_float(val):
    """Best-effort numeric coercion. VN reports use ',' thousand separators."""
    try:
        val_str = str(val).replace(',', '').strip()
        return float(val_str) if val_str not in ('', 'None') else 0.0
    except (ValueError, TypeError):
        return 0.0


def _cell(row, idx, default=''):
    """Safe positional cell access that never raises IndexError."""
    if idx is None or idx >= len(row):
        return default
    v = row[idx]
    return default if v is None else v


def _normalize_name(name):
    """Uppercase + collapse whitespace, used to match the same person across files."""
    return re.sub(r'\s+', ' ', (name or '').strip()).upper()


def _slugify_name(name):
    return re.sub(r'\s+', '_', (name or '').strip().upper())


def _extract_date_from_text(text, label):
    """Finds e.g. 'Đến ngày: 31/07/2026' anywhere in a blob of report text."""
    m = re.search(re.escape(label) + r'\s*:?\s*([\d/]{6,10})', text)
    return m.group(1) if m else None


def _to_iso_date(val):
    """
    Converts a VN 'DD/MM/YYYY' date string to ISO 'YYYY-MM-DD' for safe SQL literals.
    Postgres' default DateStyle (MDY) would misparse '30/07/2026' — returns None
    (-> SQL NULL) when the value can't be confidently parsed, rather than emitting
    something that silently swaps day/month.
    """
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', s)
    if m:
        d, mo, y = m.groups()
        try:
            return f"{y}-{int(mo):02d}-{int(d):02d}"
        except ValueError:
            return None
    if re.match(r'^\d{4}-\d{1,2}-\d{1,2}$', s):
        return s
    return None


def _load_excel_rows(file_bytes_or_path):
    """Loads a single-sheet .xls/.xlsx report (bytes or path) into a list-of-rows matrix."""
    if isinstance(file_bytes_or_path, bytes):
        if file_bytes_or_path.startswith(b'\xd0\xcf\x11\xe0'):
            wb = xlrd.open_workbook(file_contents=file_bytes_or_path)
            sheet = wb.sheet_by_index(0)
            return [sheet.row_values(i) for i in range(sheet.nrows)]
        wb = openpyxl.load_workbook(filename=io.BytesIO(file_bytes_or_path), data_only=True)
        sheet = wb.active
        return [[cell.value if cell.value is not None else '' for cell in row] for row in sheet.iter_rows()]

    if str(file_bytes_or_path).lower().endswith('.xls'):
        wb = xlrd.open_workbook(file_bytes_or_path)
        sheet = wb.sheet_by_index(0)
        return [sheet.row_values(i) for i in range(sheet.nrows)]
    wb = openpyxl.load_workbook(file_bytes_or_path, data_only=True)
    sheet = wb.active
    return [[cell.value if cell.value is not None else '' for cell in row] for row in sheet.iter_rows()]


def _header_text_blob(rows_data, n=10):
    return " | ".join(str(c) for row in rows_data[:n] for c in row if c not in (None, ''))


def _detect_excel_report_type(rows_data):
    """
    Distinguishes the two known report layouts:
      - 'order_history': "Lịch sử đặt lệnh" export (e.g. PBSV.xlsx) — one row per order,
        has 'Tiểu khoản' + 'Mua/Bán' + 'Trạng thái' columns, needs Hoàn thành filter.
      - 'trade_result':  "Kết quả khớp lệnh và phí giao dịch" style report — one row per
        sub-account/symbol with separate Mua/Bán column groups (SHL, Môi giới, CTV...).
    """
    for row in rows_data[:15]:
        cells = [str(c).strip().lower() if c is not None else '' for c in row]
        has_sub = any('tiểu khoản' in c for c in cells)
        has_side = any('mua/bán' in c for c in cells)
        has_status = any('trạng thái' in c for c in cells)
        if has_sub and has_side and has_status:
            return 'order_history'
    return 'trade_result'


# ────────────────────────────────────────────────────────────────────────────
# Parser: "Kết quả khớp lệnh và phí giao dịch" style report
# (e.g. "Bao cao giao dich *.xls" — one row per sub-account/symbol match)
#
# This report only exposes one account-level code ("Số tiểu khoản" column), so
# there is no separate parent "Số tài khoản" — the sub-account IS the account
# (parent_acc == sub_acc).
# ────────────────────────────────────────────────────────────────────────────

def _parse_trade_result_rows(rows_data, filename=""):
    tx_rows = []
    company_name = "Công ty Chứng khoán"

    for r in rows_data[:6]:
        for cell in r:
            cell_str = str(cell).strip()
            if "Công ty" in cell_str or "Chứng khoán" in cell_str:
                company_name = cell_str
                break

    header_text = _header_text_blob(rows_data)
    report_date = (_extract_date_from_text(header_text, 'Đến ngày')
                   or _extract_date_from_text(header_text, 'Từ ngày')
                   or '')

    header_row_idx = -1
    col_indices = {}

    for idx, r in enumerate(rows_data[:20]):
        r_str = [str(c).strip().lower() for c in r]
        if any("stt" in c or "shl" in c or "mã chứng khoán" in c or "mã ck" in c for c in r_str):
            header_row_idx = idx
            for c_idx, c_val in enumerate(r_str):
                if "stt" in c_val: col_indices['stt'] = c_idx
                elif "shl" in c_val or "số hiệu" in c_val: col_indices['shl'] = c_idx
                elif "tiểu khoản" in c_val or "tài khoản" in c_val or "số tk" in c_val: col_indices['sub_acc'] = c_idx
                elif "tên khách hàng" in c_val or "tên kh" in c_val: col_indices['cust_name'] = c_idx
                elif "mã chứng khoán" in c_val or "mã ck" in c_val: col_indices['symbol'] = c_idx
                elif "môi giới" in c_val or "br" in c_val: col_indices['broker'] = c_idx
                elif "ctv" in c_val: col_indices['ctv'] = c_idx
                elif "ngày" in c_val: col_indices['tx_date'] = c_idx
            break

    if header_row_idx == -1:
        header_row_idx = 9

    for i in range(header_row_idx + 1, len(rows_data)):
        r = rows_data[i]
        if not any(r):
            continue

        cell0 = str(r[0]).strip() if len(r) > 0 else ""
        if cell0.endswith('.0'):
            cell0 = cell0[:-2]

        if "tổng" in cell0.lower() or "sum" in cell0.lower():
            continue

        shl_val = str(_cell(r, col_indices.get('shl', 1))).split('.')[0]
        if not shl_val or len(shl_val) < 3:
            continue

        sub_acc = str(_cell(r, col_indices.get('sub_acc', 2))).strip() or 'TK_DEFAULT'
        cust_name = str(_cell(r, col_indices.get('cust_name', 3))).strip()
        symbol = str(_cell(r, col_indices.get('symbol', 4))).strip()

        buy_qty_matched = _safe_float(_cell(r, 6))
        buy_price_avg = _safe_float(_cell(r, 7))
        buy_val = _safe_float(_cell(r, 8))

        sell_qty_matched = _safe_float(_cell(r, 10))
        sell_price_avg = _safe_float(_cell(r, 11))
        sell_val = _safe_float(_cell(r, 12))

        fee_val = _safe_float(_cell(r, 14))
        tax_val = _safe_float(_cell(r, 16))

        broker = str(_cell(r, col_indices.get('broker', 19))).strip()
        ctv = str(_cell(r, col_indices.get('ctv', 20))).strip()
        tx_date = str(_cell(r, col_indices.get('tx_date', 21))).strip()

        tx_rows.append({
            'source': 'Excel (' + (filename or 'Uploaded') + ')',
            'company': company_name,
            'company_code': 'EXCEL_CO',
            'shl': shl_val,
            'sub_acc': sub_acc,
            'parent_acc': sub_acc,      # report has no separate account/sub-account split
            'cust_name': cust_name or 'Khách hàng',
            'symbol': symbol or 'VNINDEX',
            'buy_qty_matched': buy_qty_matched,
            'buy_price_avg': buy_price_avg,
            'buy_val': buy_val,
            'sell_qty_matched': sell_qty_matched,
            'sell_price_avg': sell_price_avg,
            'sell_val': sell_val,
            'fee_val': fee_val,
            'tax_val': tax_val,
            'broker': broker,
            'ctv': ctv,
            'tx_date': tx_date or report_date
        })

    return company_name, tx_rows


# ────────────────────────────────────────────────────────────────────────────
# Parser: "Lịch sử đặt lệnh" (order history) report — e.g. PBSV.xlsx
# One row per order attempt; only rows with Trạng thái == 'Hoàn thành' are real,
# settled transactions — 'Đã sửa' / 'Đã hủy' rows must be discarded.
#
# This report DOES expose two levels: a "Số tài khoản: XXX" account in the header,
# and a "Tiểu khoản" column per row (sub-account under that account).
# ────────────────────────────────────────────────────────────────────────────

def _parse_order_history_rows(rows_data, filename=""):
    tx_rows = []
    company_name = "Công ty Chứng khoán"

    for r in rows_data[:6]:
        for cell in r:
            cell_str = str(cell).strip()
            if "Công ty" in cell_str or "Chứng khoán" in cell_str:
                company_name = cell_str
                break

    header_text = _header_text_blob(rows_data)
    acc_match = re.search(r'Số tài khoản\s*:?\s*([A-Za-z0-9]+)', header_text)
    account_number = acc_match.group(1).strip() if acc_match else None
    report_date = (_extract_date_from_text(header_text, 'Đến ngày')
                   or _extract_date_from_text(header_text, 'Từ ngày')
                   or '')

    header_idx = -1
    col = {}
    for idx, row in enumerate(rows_data[:15]):
        cells = [str(c).strip().lower() if c is not None else '' for c in row]
        if any('tiểu khoản' in c for c in cells) and any('mua/bán' in c for c in cells):
            header_idx = idx
            for c_idx, c_val in enumerate(cells):
                if 'tiểu khoản' in c_val: col['sub'] = c_idx
                elif c_val == 'ngày': col['date'] = c_idx
                elif 'mua/bán' in c_val: col['side'] = c_idx
                elif 'mã ck' in c_val or 'mã chứng khoán' in c_val: col['symbol'] = c_idx
                elif 'thông tin giao dịch' in c_val: col['detail_start'] = c_idx
                elif 'trạng thái' in c_val: col['status'] = c_idx
                elif c_val == 'phí': col['fee'] = c_idx
                elif 'thuế tncn (quyền)' in c_val: col['tax2'] = c_idx
                elif 'thuế tncn' in c_val: col['tax'] = c_idx
                elif 'số hiệu lệnh' in c_val: col['order_id'] = c_idx
            break

    if header_idx == -1:
        return company_name, tx_rows

    # The merged "Thông tin giao dịch chứng khoán" header has its real sub-columns
    # (KL đặt / Giá đặt / KL khớp / Giá khớp / Giá trị khớp) on the NEXT row.
    data_start = header_idx + 1
    if 'detail_start' in col and header_idx + 1 < len(rows_data):
        sub_cells = [str(c).strip().lower() if c is not None else '' for c in rows_data[header_idx + 1]]
        base = col['detail_start']
        for c_idx in range(base, min(base + 6, len(sub_cells))):
            label = sub_cells[c_idx]
            if 'kl khớp' in label: col['qty_matched'] = c_idx
            elif 'giá khớp' in label: col['price_matched'] = c_idx
            elif 'giá trị khớp' in label: col['val_matched'] = c_idx
        data_start = header_idx + 2

    for i in range(data_start, len(rows_data)):
        row = rows_data[i]
        if not any(row):
            continue

        status = str(_cell(row, col.get('status'))).strip()
        if status != 'Hoàn thành':
            continue  # chỉ lấy các giao dịch đã khớp hoàn tất, bỏ 'Đã sửa' / 'Đã hủy'

        side = str(_cell(row, col.get('side', 3))).strip()
        symbol = str(_cell(row, col.get('symbol', 4))).strip()
        tieu_khoan = str(_cell(row, col.get('sub', 1))).strip()
        tx_date = str(_cell(row, col.get('date', 2))).strip()
        order_id = str(_cell(row, col.get('order_id', 19), f"OH{i}")).strip()

        qty = _safe_float(_cell(row, col.get('qty_matched')))
        price = _safe_float(_cell(row, col.get('price_matched')))
        val = _safe_float(_cell(row, col.get('val_matched')))
        fee = _safe_float(_cell(row, col.get('fee')))
        tax = _safe_float(_cell(row, col.get('tax'))) + _safe_float(_cell(row, col.get('tax2')))

        if val <= 0 and qty <= 0:
            continue

        parent_acc = account_number or (tieu_khoan or 'TK_DEFAULT')
        sub_acc = f"{parent_acc}-{tieu_khoan}" if tieu_khoan else parent_acc
        is_buy = 'mua' in side.lower()

        tx_rows.append({
            'source': 'Excel (' + (filename or 'Uploaded') + ')',
            'company': company_name,
            'company_code': 'EXCEL_CO',
            'shl': order_id,
            'sub_acc': sub_acc,
            'parent_acc': parent_acc,
            'cust_name': f"Khách hàng {parent_acc}",
            'symbol': symbol or 'VNINDEX',
            'buy_qty_matched': qty if is_buy else 0.0,
            'buy_price_avg': price if is_buy else 0.0,
            'buy_val': val if is_buy else 0.0,
            'sell_qty_matched': qty if not is_buy else 0.0,
            'sell_price_avg': price if not is_buy else 0.0,
            'sell_val': val if not is_buy else 0.0,
            'fee_val': fee,
            'tax_val': tax,
            'broker': '',
            'ctv': '',
            'tx_date': tx_date or report_date
        })

    return company_name, tx_rows


def parse_excel_data(file_bytes_or_path, filename=""):
    """
    Parses an XLS/XLSX transaction report, auto-detecting which of the two known
    report layouts it is ('trade_result' vs 'order_history') and routing to the
    matching row extractor.
    Returns: company_name, list of transaction rows
    """
    try:
        rows_data = _load_excel_rows(file_bytes_or_path)
        if not rows_data:
            return "Công ty Chứng khoán", []

        report_type = _detect_excel_report_type(rows_data)
        if report_type == 'order_history':
            return _parse_order_history_rows(rows_data, filename=filename)
        return _parse_trade_result_rows(rows_data, filename=filename)

    except Exception as e:
        print(f"Error parsing Excel file {filename}: {e}")
        return "Công ty Chứng khoán", []


def parse_pdf_data(file_bytes_or_path, filename=""):
    """
    Parses any PDF transaction report dynamically.
    Returns: company_name, list of transaction rows
    """
    if isinstance(file_bytes_or_path, bytes):
        pdf_file = io.BytesIO(file_bytes_or_path)
    else:
        pdf_file = file_bytes_or_path

    company_name = "Công ty Chứng khoán (PDF)"
    tx_rows = []

    try:
        with pdfplumber.open(pdf_file) as pdf:
            full_text = "\n".join(page.extract_text() or '' for page in pdf.pages)
            report_date = (_extract_date_from_text(full_text, 'Đến ngày')
                           or _extract_date_from_text(full_text, 'Từ ngày')
                           or '')

            for page in pdf.pages:
                text = page.extract_text()
                if not text:
                    continue
                lines = text.split('\n')
                if lines:
                    first_line = lines[0].strip()
                    if "Chứng khoán" in first_line or "Công ty" in first_line:
                        company_name = first_line

                current_manager = "Người quản lý"
                current_role = "Quản lý tài khoản"

                for line in lines:
                    line_str = line.strip()
                    if "Quản lý" in line_str or "Môi giới" in line_str:
                        mgr = re.sub(r'Quản lý.*', '', line_str).strip()
                        mgr = re.sub(r'^\d+\s*', '', mgr)
                        if mgr:
                            current_manager = mgr

                    m = re.search(r'([0-9A-Z]{6,15})\s+(.*?)\s+([A-Z0-9]{3,6})\s+([\d,.]+)\s+([\d,.]+)', line_str)
                    if m:
                        sub_acc, cust_name, symbol, buy_val_str, sell_val_str = m.groups()
                        buy_val = float(buy_val_str.replace(',', '').replace('.', ''))
                        sell_val = float(sell_val_str.replace(',', '').replace('.', ''))

                        tx_rows.append({
                            'source': 'PDF (' + (filename or 'Uploaded') + ')',
                            'company': company_name,
                            'company_code': 'PDF_CO',
                            'sub_acc': sub_acc,
                            'parent_acc': sub_acc,   # report exposes only one account-level code
                            'cust_name': cust_name,
                            'symbol': symbol,
                            'buy_val': buy_val,
                            'sell_val': sell_val,
                            'manager': current_manager,
                            'manager_role': current_role,
                            'tx_date': report_date,
                        })
    except Exception as e:
        print(f"Error parsing PDF file {filename}: {e}")

    return company_name, tx_rows


def build_relational_database(excel_tx_list=None, pdf_tx_list=None):
    """
    Transforms extracted Excel and/or PDF data into the 11 relational tables defined in
    DATA MODEL.png: cong_ty_chung_khoan, nguoi_quan_ly, chinh_sach, phan_loai_khach_hang,
    nhom_khach_hang, khach_hang, tieu_khoan, co_phieu, giao_dich, phi_gia_han, bao_cao_thu_lai.
    Works with ANY single list or combination of lists.

    Table and column names are lowercase snake_case (matching supabase_schema.sql exactly),
    e.g. 'ma_giao_dich', 'so_tai_khoan' — this is what's shown in the editable grids and
    what generate_sql_script() writes straight into INSERT statements, no translation layer.

    khach_hang is keyed by 'so_tai_khoan' (the account holder). tieu_khoan is the weak
    entity — one or more sub-accounts ('so_tieu_khoan') per Khách hàng — and giao_dich
    references tieu_khoan, not khach_hang directly.
    """
    excel_tx_list = excel_tx_list or []
    pdf_tx_list = pdf_tx_list or []
    all_tx_raw = excel_tx_list + pdf_tx_list

    # 1. cong_ty_chung_khoan
    company_map = {}
    co_id_counter = 1
    for r in all_tx_raw:
        c_name = r.get('company', 'Công ty Chứng khoán')
        if c_name not in company_map:
            c_code = "CTCK_" + str(co_id_counter)
            company_map[c_name] = {'id': co_id_counter, 'ma_dinh_danh_cong_ty': c_code, 'ten_cong_ty': c_name}
            co_id_counter += 1

    if not company_map:
        company_map['Công ty Chứng khoán Default'] = {'id': 1, 'ma_dinh_danh_cong_ty': 'CTCK_01', 'ten_cong_ty': 'Công ty Chứng khoán Default'}

    df_company = pd.DataFrame(list(company_map.values()))

    # ── Resolve a stable manager code for every PDF row by matching the person's
    #    name against managers already identified in the Excel report(s) — instead
    #    of hard-coding one specific name, this generalizes to any overlapping person.
    excel_name_to_code = {}
    for r in excel_tx_list:
        for field in ('broker', 'ctv'):
            val = r.get(field, '')
            if val and '-' in val:
                code, name = val.split('-', 1)
                excel_name_to_code[_normalize_name(name)] = code.strip()

    pdf_generated_codes = {}
    for r in pdf_tx_list:
        mgr_name = r.get('manager', 'Người quản lý')
        norm = _normalize_name(mgr_name)
        if norm in excel_name_to_code:
            r['manager_code'] = excel_name_to_code[norm]
        elif norm in pdf_generated_codes:
            r['manager_code'] = pdf_generated_codes[norm]
        else:
            code = "NQL_" + _slugify_name(mgr_name)
            pdf_generated_codes[norm] = code
            r['manager_code'] = code

    # 2. nguoi_quan_ly
    managers_dict = {}
    mgr_id_counter = 1

    for r in excel_tx_list:
        co_id = company_map.get(r.get('company'), {}).get('id', 1)
        for field, role in [('broker', 'Môi giới'), ('ctv', 'CTV')]:
            val = r.get(field, '')
            if val:
                parts = val.split('-')
                code = parts[0].strip()
                name = parts[1].strip() if len(parts) > 1 else parts[0].strip()
                if code not in managers_dict:
                    managers_dict[code] = {
                        'id': mgr_id_counter,
                        'ma_nguoi_quan_ly_ctv': code,
                        'ten_nguoi_quan_ly_ctv': name,
                        'loai_nguoi_quan_ly': role,
                        'ma_cong_ty_chung_khoan': co_id,
                        'tinh_trang_hoat_dong': 1
                    }
                    mgr_id_counter += 1

    for r in pdf_tx_list:
        co_id = company_map.get(r.get('company'), {}).get('id', 1)
        mgr_name = r.get('manager', 'Người quản lý')
        code = r.get('manager_code') or ("NQL_" + _slugify_name(mgr_name))
        if code not in managers_dict:
            managers_dict[code] = {
                'id': mgr_id_counter,
                'ma_nguoi_quan_ly_ctv': code,
                'ten_nguoi_quan_ly_ctv': mgr_name,
                'loai_nguoi_quan_ly': r.get('manager_role', 'Quản lý tài khoản'),
                'ma_cong_ty_chung_khoan': co_id,
                'tinh_trang_hoat_dong': 1
            }
            mgr_id_counter += 1

    if not managers_dict:
        managers_dict['NQL_DEF'] = {
            'id': 1, 'ma_nguoi_quan_ly_ctv': 'NQL_DEF', 'ten_nguoi_quan_ly_ctv': 'Người quản lý hệ thống',
            'loai_nguoi_quan_ly': 'Quản lý', 'ma_cong_ty_chung_khoan': 1, 'tinh_trang_hoat_dong': 1
        }

    df_manager = pd.DataFrame(list(managers_dict.values()))

    # 3. phan_loai_khach_hang & nhom_khach_hang & chinh_sach
    df_chinh_sach = pd.DataFrame([
        {
            'ma_chinh_sach': 'CS01',
            'ten_chinh_sach': 'Chính sách giao dịch chuẩn',
            'lai_suat': 0.105,
            'phi_giao_dich': 0.00075,
            'phi_ung_truoc': 0.0003,
            'phi_gia_han': 0.0005,
            'lai_gia_han': 0.0002,
            'thoi_han': 90,
            'han_muc_tong': 50000000000.0,
            'ty_le_vay': 0.5
        }
    ])

    df_phan_loai_kh = pd.DataFrame([
        {
            'ma_loai_khach_hang': 'LKH01',
            'ten_loai_khach_hang': 'Khách hàng cá nhân',
            'phan_loai': 'Cá nhân',
            'mo_ta': 'Tài khoản giao dịch cá nhân',
            'ma_chinh_sach': 'CS01'
        },
        {
            'ma_loai_khach_hang': 'LKH02',
            'ten_loai_khach_hang': 'Khách hàng tổ chức',
            'phan_loai': 'Tổ chức',
            'mo_ta': 'Doanh nghiệp / Quỹ đầu tư',
            'ma_chinh_sach': 'CS01'
        }
    ])

    df_nhom_kh = pd.DataFrame([
        {
            'ma_nhom_khach_hang': 'NKH01',
            'ten_nhom_khach_hang': 'Nhóm giao dịch thường',
            'phan_nhom': 'Tiêu chuẩn',
            'ma_chinh_sach': 'CS01',
            'mo_ta': 'Nhóm khách hàng cơ bản'
        }
    ])

    # 4. khach_hang (PK = so_tai_khoan) — keyed by parent_acc, so several tiểu khoản
    #    (sub_acc) that share the same parent account collapse into ONE customer row.
    customers_dict = {}
    for r in all_tx_raw:
        sub_acc = r.get('sub_acc', 'TK_DEFAULT')
        parent_acc = r.get('parent_acc') or sub_acc
        if parent_acc not in customers_dict:
            co_id = company_map.get(r.get('company'), {}).get('id', 1)
            mgr_code = 'NQL_DEF'
            if r.get('ctv'): mgr_code = r['ctv'].split('-')[0]
            elif r.get('broker'): mgr_code = r['broker'].split('-')[0]
            elif r.get('manager_code'): mgr_code = r['manager_code']

            is_org = is_company_or_organization(r.get('cust_name', ''))
            cust_type = 'LKH02' if is_org else 'LKH01'

            customers_dict[parent_acc] = {
                'so_tai_khoan': parent_acc,
                'ten_khach_hang': format_customer_name(r.get('cust_name', 'Khách hàng')),
                'ma_cong_ty_chung_khoan': co_id,
                'ma_loai_khach_hang': cust_type,
                'ma_nhom_khach_hang': 'NKH01',
                'ma_nguoi_quan_ly': mgr_code,
                'nav': 0.0,
                'du_no_goc': 0.0,
                'du_no_lai': 0.0,
                'ngay_toi_han_gan_nhat': None,
                'ghi_chu': f"Imported from {r.get('source', 'Report')}",
                'tinh_trang_hoat_dong': 1,
                'tong_du_no': 0.0
            }

    df_customer = pd.DataFrame(list(customers_dict.values()))

    # 5. tieu_khoan (weak entity, PK = so_tieu_khoan + so_tai_khoan)
    tieu_khoan_dict = {}
    for r in all_tx_raw:
        sub_acc = r.get('sub_acc', 'TK_DEFAULT')
        parent_acc = r.get('parent_acc') or sub_acc
        if sub_acc not in tieu_khoan_dict:
            tieu_khoan_dict[sub_acc] = {
                'so_tieu_khoan': sub_acc,
                'so_tai_khoan': parent_acc
            }
    df_tieu_khoan = pd.DataFrame(list(tieu_khoan_dict.values()))

    # 6. co_phieu
    stocks_dict = {}
    for r in all_tx_raw:
        sym = r.get('symbol', 'VNINDEX')
        if sym and sym not in stocks_dict:
            price = r.get('buy_price_avg', 0.0) or r.get('sell_price_avg', 0.0) or 0.0
            stocks_dict[sym] = {
                'ma_co_phieu': sym,
                'ten_doanh_nghiep': f"CTCP {sym}",
                'gia_mo_cua_ngay_giao_dich_gan_nhat': price,
                'gia_dong_cua_ngay_giao_dich_gan_nhat': price
            }
    df_stock = pd.DataFrame(list(stocks_dict.values()))

    # 7. giao_dich (FK -> tieu_khoan.so_tieu_khoan, not khach_hang directly)
    tx_list = []

    # Excel transactions (both "trade_result" and "order_history" rows share this shape)
    for r in excel_tx_list:
        mgr_code = 'NQL_DEF'
        if r.get('ctv'): mgr_code = r['ctv'].split('-')[0]
        elif r.get('broker'): mgr_code = r['broker'].split('-')[0]

        if r.get('buy_val', 0) > 0 or r.get('buy_qty_matched', 0) > 0:
            tx_list.append({
                'ma_giao_dich': r['shl'],
                'so_tieu_khoan': r['sub_acc'],
                'ma_nguoi_quan_ly': mgr_code,
                'gia_tri_giao_dich': r['buy_val'],
                'giao_dich_mua_ban': 1,
                'ma_cp': r['symbol'],
                'thue_ban': 0.0,
                'ngay_giao_dich': r.get('tx_date') or '',
                'phi_net': r.get('fee_val', 0.0),
                'khoi_luong_giao_dich': r.get('buy_qty_matched', 0.0),
                'gia_giao_dich': r.get('buy_price_avg', 0.0)
            })

        if r.get('sell_val', 0) > 0 or r.get('sell_qty_matched', 0) > 0:
            tx_list.append({
                'ma_giao_dich': r['shl'],
                'so_tieu_khoan': r['sub_acc'],
                'ma_nguoi_quan_ly': mgr_code,
                'gia_tri_giao_dich': r['sell_val'],
                'giao_dich_mua_ban': 2,
                'ma_cp': r['symbol'],
                'thue_ban': r.get('tax_val', 0.0),
                'ngay_giao_dich': r.get('tx_date') or '',
                'phi_net': r.get('fee_val', 0.0),
                'khoi_luong_giao_dich': r.get('sell_qty_matched', 0.0),
                'gia_giao_dich': r.get('sell_price_avg', 0.0)
            })

    # PDF transactions (generate sequential ID: sub_acc + symbol + index)
    pdf_seq = len(tx_list)  # Continue sequence after excel transactions
    for idx, r in enumerate(pdf_tx_list, start=1):
        mgr_code = r.get('manager_code', 'NQL_DEF')
        sub_acc = r['sub_acc']
        symbol = r['symbol']
        tx_date = r.get('tx_date') or ''

        if r.get('buy_val', 0) > 0:
            pdf_seq += 1
            tx_list.append({
                'ma_giao_dich': f"{sub_acc}{symbol}{str(pdf_seq).zfill(4)}",
                'so_tieu_khoan': sub_acc,
                'ma_nguoi_quan_ly': mgr_code,
                'gia_tri_giao_dich': r['buy_val'],
                'giao_dich_mua_ban': 1,
                'ma_cp': symbol,
                'thue_ban': 0.0,
                'ngay_giao_dich': tx_date,
                'phi_net': r['buy_val'] * 0.00075,
                'khoi_luong_giao_dich': 0.0,
                'gia_giao_dich': 0.0
            })

        if r.get('sell_val', 0) > 0:
            pdf_seq += 1
            tx_list.append({
                'ma_giao_dich': f"{sub_acc}{symbol}{str(pdf_seq).zfill(4)}",
                'so_tieu_khoan': sub_acc,
                'ma_nguoi_quan_ly': mgr_code,
                'gia_tri_giao_dich': r['sell_val'],
                'giao_dich_mua_ban': 2,
                'ma_cp': symbol,
                'thue_ban': r['sell_val'] * 0.001,
                'ngay_giao_dich': tx_date,
                'phi_net': r['sell_val'] * 0.00075,
                'khoi_luong_giao_dich': 0.0,
                'gia_giao_dich': 0.0
            })

    df_giao_dich = pd.DataFrame(tx_list) if tx_list else pd.DataFrame(columns=[
        'ma_giao_dich', 'so_tieu_khoan', 'ma_nguoi_quan_ly', 'gia_tri_giao_dich',
        'giao_dich_mua_ban', 'ma_cp', 'thue_ban', 'ngay_giao_dich',
        'phi_net', 'khoi_luong_giao_dich', 'gia_giao_dich'
    ])

    # 8. phi_gia_han & 9. bao_cao_thu_lai
    # No source report supplies phí gia hạn / lãi thu data — leave these empty with
    # the correct columns rather than fabricating a fake all-zero placeholder row
    # that the user would just have to notice and delete by hand.
    df_phi_gia_han = pd.DataFrame(columns=[
        'id', 'ngay', 'ma_khach_hang', 'phi_gia_han_du_thu', 'phi_gia_han_thuc_thu', 'lai'
    ])
    df_bao_cao_thu_lai = pd.DataFrame(columns=[
        'id', 'ngay_thu_lai', 'ma_khach_hang', 'lai_vay', 'lai_ung_truoc'
    ])

    return {
        'cong_ty_chung_khoan': df_company,
        'nguoi_quan_ly': df_manager,
        'chinh_sach': df_chinh_sach,
        'phan_loai_khach_hang': df_phan_loai_kh,
        'nhom_khach_hang': df_nhom_kh,
        'khach_hang': df_customer,
        'tieu_khoan': df_tieu_khoan,
        'co_phieu': df_stock,
        'giao_dich': df_giao_dich,
        'phi_gia_han': df_phi_gia_han,
        'bao_cao_thu_lai': df_bao_cao_thu_lai
    }


def build_master_flat_table(db_tables):
    """
    Creates a unified master flat dataset combining all attributes from Data Model.
    Handles both ETL-generated tables and PBSV multi-sheet workbook column naming conventions.
    Only includes columns that actually have data — never fills in values from wrong columns.
    """
    df_gd  = db_tables.get('giao_dich', pd.DataFrame()).copy()
    df_kh  = db_tables.get('khach_hang', pd.DataFrame()).copy()
    df_tk  = db_tables.get('tieu_khoan', pd.DataFrame()).copy()
    df_nql = db_tables.get('nguoi_quan_ly', pd.DataFrame()).copy()
    df_co  = db_tables.get('cong_ty_chung_khoan', pd.DataFrame()).copy()
    df_cp  = db_tables.get('co_phieu', pd.DataFrame()).copy()

    if df_gd.empty:
        return pd.DataFrame()

    # ── Detect column naming convention ──────────────────────────────
    # ETL-generated:   'so_tieu_khoan', 'ma_nguoi_quan_ly', 'ma_cp', ...
    # PBSV multi-sheet: 'Mã KH (FK)', 'Mã người QL (FK)', 'Mã CP (FK)', ... (raw headers
    # from the uploaded workbook itself — outside our control, so left un-translated)
    is_pbsv = 'Mã KH (FK)' in df_gd.columns

    if is_pbsv:
        # ── Rename PBSV columns → canonical snake_case names for joining ──
        gd_rename = {
            'Mã KH (FK)':        'ma_khach_hang',
            'Tên KH':            'ten_khach_hang__gd',  # keep separately
            'Mã người QL (FK)':  '_mgr_fk',
            'Tên người QL':      'ten_nguoi_quan_ly__gd',
            'Mua/Bán (1:Mua, 2:Bán)': 'giao_dich_mua_ban',
            'Mã CP (FK)':        'ma_cp',
            'Tên DN (CP)':       'ten_doanh_nghiep__gd',
            'Ngày giao dịch':    'ngay_giao_dich',
            'Khối lượng GD':     'khoi_luong_giao_dich',
            'Giá GD':            'gia_giao_dich',
        }
        df_gd.rename(columns={k: v for k, v in gd_rename.items() if k in df_gd.columns}, inplace=True)

        kh_rename = {
            'Mã KH':                         'ma_khach_hang',
            'Mã Cty CK (FK)':                'ma_cong_ty_chung_khoan',
            'Mã loại KH (FK)':               'ma_loai_khach_hang',
            'Mã nhóm KH (FK)':               'ma_nhom_khach_hang',
            'Mã người QL (FK)':              '_kh_mgr_fk',
            'Tình trạng hoạt động (1/0)':   'tinh_trang_hoat_dong',
        }
        df_kh.rename(columns={k: v for k, v in kh_rename.items() if k in df_kh.columns}, inplace=True)

        nql_rename = {
            'Mã người QL/CTV': 'ma_nguoi_quan_ly_ctv',
            'Tên người QL/CTV': 'ten_nguoi_quan_ly_ctv',
            'Loại (Quản lý/CTV)': 'loai_nguoi_quan_ly',
        }
        df_nql.rename(columns={k: v for k, v in nql_rename.items() if k in df_nql.columns}, inplace=True)

        cp_rename = {
            'Mã cổ phiếu': 'ma_co_phieu',
            'Giá mở cửa (gần nhất)': 'gia_mo_cua_ngay_giao_dich_gan_nhat',
            'Giá đóng cửa (gần nhất)': 'gia_dong_cua_ngay_giao_dich_gan_nhat',
        }
        df_cp.rename(columns={k: v for k, v in cp_rename.items() if k in df_cp.columns}, inplace=True)

        # Build ma_nguoi_quan_ly in giao_dich from NQL lookup
        if '_mgr_fk' in df_gd.columns and 'id' in df_nql.columns:
            nql_lookup = df_nql[['id', 'ma_nguoi_quan_ly_ctv']].copy()
            nql_lookup['id'] = pd.to_numeric(nql_lookup['id'], errors='coerce')
            df_gd['_mgr_fk'] = pd.to_numeric(df_gd['_mgr_fk'], errors='coerce')
            df_gd = pd.merge(df_gd, nql_lookup, left_on='_mgr_fk', right_on='id', how='left')
            df_gd.rename(columns={'ma_nguoi_quan_ly_ctv': 'ma_nguoi_quan_ly'}, inplace=True)
            df_gd.drop(columns=['_mgr_fk', 'id'], errors='ignore', inplace=True)

        # ── Merge giao_dich ← khach_hang (legacy PBSV convention: direct FK) ──
        if 'ma_khach_hang' in df_gd.columns and 'ma_khach_hang' in df_kh.columns:
            kh_cols_to_keep = [c for c in df_kh.columns
                               if c not in df_gd.columns or c == 'ma_khach_hang']
            master_df = pd.merge(df_gd, df_kh[kh_cols_to_keep], on='ma_khach_hang', how='left')
        else:
            master_df = df_gd.copy()
    else:
        # ── ETL convention: giao_dich.so_tieu_khoan -> tieu_khoan -> khach_hang.so_tai_khoan ──
        if 'so_tieu_khoan' in df_gd.columns and 'so_tieu_khoan' in df_tk.columns:
            master_df = pd.merge(df_gd, df_tk, on='so_tieu_khoan', how='left')
        else:
            master_df = df_gd.copy()

        if 'so_tai_khoan' in master_df.columns and 'so_tai_khoan' in df_kh.columns:
            kh_cols_to_keep = [c for c in df_kh.columns
                               if c not in master_df.columns or c == 'so_tai_khoan']
            master_df = pd.merge(master_df, df_kh[kh_cols_to_keep], on='so_tai_khoan', how='left')

    # ── Merge ← nguoi_quan_ly ────────────────────────────────────────
    if 'ma_nguoi_quan_ly' in master_df.columns and 'ma_nguoi_quan_ly_ctv' in df_nql.columns:
        nql_cols = [c for c in df_nql.columns
                    if c not in master_df.columns or c == 'ma_nguoi_quan_ly_ctv']
        master_df = pd.merge(master_df, df_nql[nql_cols],
                             left_on='ma_nguoi_quan_ly', right_on='ma_nguoi_quan_ly_ctv',
                             how='left', suffixes=('', '__nql'))

    # ── Merge ← cong_ty_chung_khoan ─────────────────────────────────
    co_id_col = None
    if 'ma_cong_ty_chung_khoan' in master_df.columns:
        co_id_col = 'ma_cong_ty_chung_khoan'
    if co_id_col and 'id' in df_co.columns:
        co_cols = [c for c in df_co.columns
                   if c not in master_df.columns or c == 'id']
        master_df = pd.merge(master_df, df_co[co_cols],
                             left_on=co_id_col, right_on='id',
                             how='left', suffixes=('', '__co'))

    # ── Merge ← co_phieu ─────────────────────────────────────────────
    if 'ma_cp' in master_df.columns and 'ma_co_phieu' in df_cp.columns:
        cp_cols = [c for c in df_cp.columns
                   if c not in master_df.columns or c == 'ma_co_phieu']
        master_df = pd.merge(master_df, df_cp[cp_cols],
                             left_on='ma_cp', right_on='ma_co_phieu',
                             how='left', suffixes=('', '__cp'))

    # ── Drop internal helper / duplicate columns ──────────────────────
    always_drop = {
        '_mgr_fk', '_kh_mgr_fk',
        'Tên Cty CK', 'Tên công ty CK', 'MucLuc',
        'id',                                        # numeric PK used only for join
        'ma_co_phieu',                               # duplicate of 'ma_cp'
        'Mã Cty CK (FK)',                            # raw FK kept after rename
        'ma_cong_ty_chung_khoan',                    # numeric FK, ma_dinh_danh_cong_ty is cleaner
        'Tên người QL',                              # duplicate of 'ten_nguoi_quan_ly_ctv'
        'Tên loại KH',                               # short alias, ma_loai_khach_hang is sufficient
        'Tên nhóm KH',                               # short alias, ma_nhom_khach_hang is sufficient
        'Tình trạng hoạt động (1:có, 0:không)',      # already renamed to canonical
        'ma_nguoi_quan_ly',                          # keep only ma_nguoi_quan_ly_ctv (same value)
    }
    drop_cols = [c for c in master_df.columns
                 if c.endswith(('__gd', '__kh', '__nql', '__co', '__cp')) or c in always_drop]
    master_df.drop(columns=drop_cols, errors='ignore', inplace=True)
    dup_cols = [c for c in master_df.columns if c.endswith('_x') or c.endswith('_y')]
    master_df.drop(columns=dup_cols, errors='ignore', inplace=True)
    master_df = master_df.loc[:, master_df.notna().any(axis=0)]

    # ── Preferred column order (only include columns that exist) ──────
    preferred = [
        'ma_giao_dich', 'ngay_giao_dich', 'giao_dich_mua_ban',
        'so_tieu_khoan', 'so_tai_khoan', 'ten_khach_hang',
        'ma_cp', 'ten_doanh_nghiep', 'gia_tri_giao_dich',
        'khoi_luong_giao_dich', 'gia_giao_dich', 'thue_ban', 'phi_net',
        'ma_nguoi_quan_ly_ctv', 'ten_nguoi_quan_ly_ctv',
        'loai_nguoi_quan_ly',
        'ma_dinh_danh_cong_ty', 'ten_cong_ty',
        'ma_loai_khach_hang', 'ma_nhom_khach_hang',
        'nav', 'du_no_goc', 'du_no_lai', 'ngay_toi_han_gan_nhat',
        'ghi_chu', 'tinh_trang_hoat_dong',
        'tong_du_no',
        'gia_mo_cua_ngay_giao_dich_gan_nhat', 'gia_dong_cua_ngay_giao_dich_gan_nhat',
    ]
    final_cols = [c for c in preferred if c in master_df.columns
                  and master_df[c].notna().any()]
    extra = [c for c in master_df.columns if c not in final_cols]
    return master_df[final_cols + extra].reset_index(drop=True)


# ────────────────────────────────────────────────────────────────────────────
# SQL generation — table/column names in db_tables are already the exact
# snake_case identifiers used in supabase_schema.sql, so this just emits
# CREATE TABLE IF NOT EXISTS + INSERT statements directly from them.
# ────────────────────────────────────────────────────────────────────────────

# Each entry: (table_name, CREATE TABLE ddl, {date columns}, needs_identity_override)
_TABLE_DEFS = [
    (
        'cong_ty_chung_khoan',
        "CREATE TABLE IF NOT EXISTS cong_ty_chung_khoan ("
        "id BIGINT GENERATED ALWAYS AS IDENTITY PRIMARY KEY, "
        "ma_dinh_danh_cong_ty VARCHAR(50), ten_cong_ty VARCHAR(255));",
        set(), True,
    ),
    (
        'chinh_sach',
        "CREATE TABLE IF NOT EXISTS chinh_sach ("
        "ma_chinh_sach VARCHAR(50) PRIMARY KEY, ten_chinh_sach VARCHAR(255), "
        "lai_suat NUMERIC(10,4), phi_giao_dich NUMERIC(10,4), phi_ung_truoc NUMERIC(10,4), "
        "phi_gia_han NUMERIC(10,4), lai_gia_han NUMERIC(10,4), thoi_han INT, "
        "han_muc_tong NUMERIC(18,2), ty_le_vay NUMERIC(10,4));",
        set(), False,
    ),
    (
        'nguoi_quan_ly',
        "CREATE TABLE IF NOT EXISTS nguoi_quan_ly ("
        "id BIGINT GENERATED ALWAYS AS IDENTITY PRIMARY KEY, ma_nguoi_quan_ly_ctv VARCHAR(50) UNIQUE, "
        "ten_nguoi_quan_ly_ctv VARCHAR(255), loai_nguoi_quan_ly VARCHAR(50), "
        "ma_cong_ty_chung_khoan BIGINT REFERENCES cong_ty_chung_khoan(id), "
        "tinh_trang_hoat_dong SMALLINT DEFAULT 1);",
        set(), True,
    ),
    (
        'phan_loai_khach_hang',
        "CREATE TABLE IF NOT EXISTS phan_loai_khach_hang ("
        "ma_loai_khach_hang VARCHAR(50) PRIMARY KEY, ten_loai_khach_hang VARCHAR(255), "
        "phan_loai VARCHAR(100), mo_ta TEXT, ma_chinh_sach VARCHAR(50) REFERENCES chinh_sach(ma_chinh_sach));",
        set(), False,
    ),
    (
        'nhom_khach_hang',
        "CREATE TABLE IF NOT EXISTS nhom_khach_hang ("
        "ma_nhom_khach_hang VARCHAR(50) PRIMARY KEY, ten_nhom_khach_hang VARCHAR(255), "
        "phan_nhom VARCHAR(100), ma_chinh_sach VARCHAR(50) REFERENCES chinh_sach(ma_chinh_sach), mo_ta TEXT);",
        set(), False,
    ),
    (
        'co_phieu',
        "CREATE TABLE IF NOT EXISTS co_phieu ("
        "ma_co_phieu VARCHAR(20) PRIMARY KEY, ten_doanh_nghiep VARCHAR(255), "
        "gia_mo_cua_ngay_giao_dich_gan_nhat NUMERIC(18,2), gia_dong_cua_ngay_giao_dich_gan_nhat NUMERIC(18,2));",
        set(), False,
    ),
    (
        'khach_hang',
        "CREATE TABLE IF NOT EXISTS khach_hang ("
        "so_tai_khoan VARCHAR(50) PRIMARY KEY, ten_khach_hang VARCHAR(255), "
        "ma_cong_ty_chung_khoan BIGINT REFERENCES cong_ty_chung_khoan(id), "
        "ma_loai_khach_hang VARCHAR(50) REFERENCES phan_loai_khach_hang(ma_loai_khach_hang), "
        "ma_nhom_khach_hang VARCHAR(50) REFERENCES nhom_khach_hang(ma_nhom_khach_hang), "
        "ma_nguoi_quan_ly VARCHAR(50) REFERENCES nguoi_quan_ly(ma_nguoi_quan_ly_ctv), "
        "nav NUMERIC(18,2) DEFAULT 0, du_no_goc NUMERIC(18,2) DEFAULT 0, du_no_lai NUMERIC(18,2) DEFAULT 0, "
        "ngay_toi_han_gan_nhat DATE, ghi_chu TEXT, tinh_trang_hoat_dong SMALLINT DEFAULT 1, "
        "tong_du_no NUMERIC(18,2) DEFAULT 0);",
        {'ngay_toi_han_gan_nhat'}, False,
    ),
    (
        'tieu_khoan',
        "CREATE TABLE IF NOT EXISTS tieu_khoan ("
        "so_tieu_khoan VARCHAR(50) NOT NULL, "
        "so_tai_khoan VARCHAR(50) NOT NULL REFERENCES khach_hang(so_tai_khoan), "
        "PRIMARY KEY (so_tieu_khoan, so_tai_khoan), UNIQUE (so_tieu_khoan));",
        set(), False,
    ),
    (
        'giao_dich',
        "CREATE TABLE IF NOT EXISTS giao_dich ("
        "ma_giao_dich VARCHAR(100) PRIMARY KEY, so_tieu_khoan VARCHAR(50) REFERENCES tieu_khoan(so_tieu_khoan), "
        "ma_nguoi_quan_ly VARCHAR(50) REFERENCES nguoi_quan_ly(ma_nguoi_quan_ly_ctv), "
        "gia_tri_giao_dich NUMERIC(18,2), giao_dich_mua_ban SMALLINT, "
        "ma_cp VARCHAR(20) REFERENCES co_phieu(ma_co_phieu), thue_ban NUMERIC(18,2), "
        "ngay_giao_dich DATE, phi_net NUMERIC(18,2), khoi_luong_giao_dich NUMERIC(18,2), "
        "gia_giao_dich NUMERIC(18,2));",
        {'ngay_giao_dich'}, False,
    ),
    (
        'phi_gia_han',
        "CREATE TABLE IF NOT EXISTS phi_gia_han ("
        "id BIGINT GENERATED ALWAYS AS IDENTITY PRIMARY KEY, ngay DATE, "
        "ma_khach_hang VARCHAR(50) REFERENCES khach_hang(so_tai_khoan), "
        "phi_gia_han_du_thu NUMERIC(18,2), phi_gia_han_thuc_thu NUMERIC(18,2), lai NUMERIC(18,2));",
        {'ngay'}, True,
    ),
    (
        'bao_cao_thu_lai',
        "CREATE TABLE IF NOT EXISTS bao_cao_thu_lai ("
        "id BIGINT GENERATED ALWAYS AS IDENTITY PRIMARY KEY, ngay_thu_lai DATE, "
        "ma_khach_hang VARCHAR(50) REFERENCES khach_hang(so_tai_khoan), "
        "lai_vay NUMERIC(18,2), lai_ung_truoc NUMERIC(18,2));",
        {'ngay_thu_lai'}, True,
    ),
]


def generate_sql_script(db_tables):
    """
    Generates SQL DDL + DML statements for PostgreSQL / Supabase, matching the
    table/column names used in supabase_schema.sql exactly (so the INSERTs work
    whether run against these CREATE TABLE IF NOT EXISTS statements or against a
    database already provisioned from supabase_schema.sql).
    """
    sql_lines = ["-- SQL Script generated by Streamlit Data Engineering Tool\n"]

    for tbl_name, ddl, date_cols, needs_override in _TABLE_DEFS:
        sql_lines.append(f"-- Table: {tbl_name}\n{ddl}")
        df = db_tables.get(tbl_name)
        if df is not None and not df.empty:
            sql_cols = list(df.columns)
            col_str = ", ".join(sql_cols)
            override_clause = "OVERRIDING SYSTEM VALUE " if needs_override and 'id' in sql_cols else ""

            for _, row in df.iterrows():
                vals = []
                for col, v in zip(df.columns, row.values):
                    if pd.isna(v) or v is None or v == '':
                        vals.append("NULL")
                    elif col in date_cols:
                        iso = _to_iso_date(v)
                        vals.append(f"'{iso}'" if iso else "NULL")
                    elif isinstance(v, (int, float)):
                        vals.append(str(v))
                    else:
                        clean_v = str(v).replace("'", "''")
                        vals.append(f"'{clean_v}'")
                sql_lines.append(
                    f"INSERT INTO {tbl_name} ({col_str}) {override_clause}"
                    f"VALUES ({', '.join(vals)}) ON CONFLICT DO NOTHING;"
                )
        sql_lines.append("\n")

    return "\n".join(sql_lines)
